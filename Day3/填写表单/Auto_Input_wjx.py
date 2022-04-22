from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import time

# 1、初始化浏览器驱动器
def init_driver():
    path = 'F:\Python自动化办公训练营\Day3\填写表单\chromedriver.exe'
    chrome_options = Options()
    # 1、 添加user-agent
    chrome_options.add_argument(
        'user-agent="Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.835.163 Safari/535.1"')
    # 2、把配置属性传递给驱动程序、并制定驱动程序的路径
    driver = webdriver.Chrome(executable_path=path, options=chrome_options)
    # 3、隐式等待（给每个网页的元素5秒钟的时间，5秒过后不出现、定位网页元素不存在、并进入异常处理）
    driver.implicitly_wait(5)
    return driver

# 2、读取Excle
def read_excel():

    wb = load_workbook("用户表.xlsx")

    ws = wb.active
    list_users = []
    for row in ws.iter_rows(min_row=2, max_row=4, min_col=1, max_col=4, values_only=True):
        list_users.append(row)

    return list_users

# 1+1 =2
# 2 *3 = 6
#
# 6+1 = 7


# 3、填写表单
def input_form(user):

    # 打开问卷星系统
    driver.get('https://www.wjx.top/jq/81172286.aspx')
    driver.find_element_by_name('q1').send_keys(user[0])
    if user[1] =='男':
        # 点击男性
        driver.find_elements_by_class_name('jqRadio')[0].click()
    else:
        # 点击女性
        driver.find_elements_by_class_name('jqRadio')[1].click()

    # 手机号
    driver.find_element_by_name('q3').send_keys(user[2])
    # 建议
    driver.find_element_by_name('q4').send_keys(user[3])
    # 提交
    driver.find_element_by_id('submit_button').click()
    time.sleep(3)


if __name__ == '__main__':

    # 初始化谷歌浏览器驱动器
    driver = init_driver()

    # 读取Excel 文件，把每行数据放到列表里。
    list_users = read_excel()

    # 遍历列表，取出用户信息
    for user in list_users:
        # 执行填写表单操作
        input_form(user)
    # 最后退出浏览器
    driver.quit()














